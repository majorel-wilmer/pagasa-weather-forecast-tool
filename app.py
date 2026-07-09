from __future__ import annotations

from datetime import datetime, timedelta
from calendar import monthrange
from io import BytesIO
from pathlib import Path
import hashlib
import hmac
import json
import os
import re
import secrets
import time

import requests
from bs4 import BeautifulSoup
from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"

# Tropical-cyclone advisories are not part of Open-Meteo's product, so the storm
# monitor page still reads PAGASA's official daily bulletin for that one feed.
PAGASA_DAILY_URL = "https://www.pagasa.dost.gov.ph/weather"
PANAHON_URL = "https://www.panahon.gov.ph/"

# Everything else (the 5-day matrix, rain timing, historical comparison) is
# sourced from Open-Meteo: https://open-meteo.com/
OPEN_METEO_FORECAST = "https://api.open-meteo.com/v1/forecast"
OPEN_METEO_ARCHIVE = "https://archive-api.open-meteo.com/v1/archive"
OPEN_METEO_ATTRIBUTION = "https://open-meteo.com/"

LOCAL_OVERRIDES = ROOT / "data" / "overrides.json"
OVERRIDE_PREFIX = "pagasa-weather-overrides/"
_override_cache = {"loaded_at": 0.0, "data": {}}

# Each site is forecast individually from its own coordinates, so no site has to
# borrow another city's numbers the way the old PAGASA-table mapping required.
SITES = [
    ("LUZON", "Alabang"),
    ("", "Antipolo"),
    ("", "Baguio"),
    ("", "Clark"),
    ("", "Laoag"),
    ("", "Metro Manila"),
    ("", "Molino"),
    ("VISAYAS", "Bacolod"),
    ("", "Cebu"),
    ("MINDANAO", "CDO"),
    ("", "Davao"),
    ("", "GenSan"),
]

SITE_COORDS = {
    "Alabang": (14.419, 121.044), "Antipolo": (14.586, 121.176),
    "Baguio": (16.402, 120.596), "Clark": (15.186, 120.560),
    "Laoag": (18.198, 120.594), "Metro Manila": (14.599, 120.984),
    "Molino": (14.396, 120.974), "Bacolod": (10.676, 122.951),
    "Cebu": (10.315, 123.885), "CDO": (8.454, 124.632),
    "Davao": (7.190, 125.455), "GenSan": (6.116, 125.171),
}

# WMO weather codes used by Open-Meteo's "weathercode" field.
WEATHER_CODE_TEXT = {
    0: "Clear sky", 1: "Mainly clear", 2: "Partly cloudy", 3: "Overcast",
    45: "Fog", 48: "Depositing rime fog",
    51: "Light drizzle", 53: "Moderate drizzle", 55: "Dense drizzle",
    56: "Light freezing drizzle", 57: "Dense freezing drizzle",
    61: "Slight rain", 63: "Moderate rain", 65: "Heavy rain",
    66: "Light freezing rain", 67: "Heavy freezing rain",
    71: "Slight snow fall", 73: "Moderate snow fall", 75: "Heavy snow fall", 77: "Snow grains",
    80: "Slight rain showers", 81: "Moderate rain showers", 82: "Violent rain showers",
    85: "Slight snow showers", 86: "Heavy snow showers",
    95: "Thunderstorm", 96: "Thunderstorm with slight hail", 99: "Thunderstorm with heavy hail",
}
THUNDER_CODES = {95, 96, 99}
RAIN_HOUR_PROB_THRESHOLD = 40  # % chance used to flag an hour as "rain expected"
RAIN_HOUR_PRECIP_THRESHOLD = 0.1  # mm in the hour

# Narrative-sentence thresholds (used for the human-readable Excel cells).
BROAD_PROB_THRESHOLD = 30     # % chance used to flag an hour as "some rain risk"
BROAD_PRECIP_THRESHOLD = 0.05  # mm in the hour
HEAVY_HOUR_PRECIP = 2.5        # mm/hr that counts as a standout "peak" hour
FULL_DAY_HOURS = 21           # flagged hours at/above this = "throughout the day"

# General sky-condition phrasing per WMO weather code (used as the lead-in for
# the plain-language forecast sentence, e.g. "Partly cloudy skies with a high
# chance of light rain from 7 AM to 11 PM.").
SKY_PHRASES = {
    0: "Clear", 1: "Clear to partly cloudy", 2: "Partly cloudy", 3: "Overcast",
    45: "Foggy", 48: "Foggy",
    51: "Partly cloudy", 53: "Partly cloudy", 55: "Cloudy",
    56: "Partly cloudy", 57: "Cloudy",
    61: "Partly cloudy", 63: "Cloudy", 65: "Widespread cloudy",
    66: "Cloudy", 67: "Widespread cloudy",
    71: "Overcast", 73: "Overcast", 75: "Overcast", 77: "Overcast",
    80: "Partly cloudy", 81: "Cloudy", 82: "Widespread cloudy",
    85: "Overcast", 86: "Overcast",
    95: "Overcast", 96: "Overcast", 99: "Overcast",
}
CHANCE_PHRASES = {
    "low": "a low chance", "medium": "a medium chance",
    "medium_high": "a medium to high chance", "high": "a high chance",
}

app = FastAPI(title="Open-Meteo 5-Day Weather Tool")
app.mount("/static", StaticFiles(directory=STATIC), name="static")


class LoginPayload(BaseModel):
    password: str


class OverridePayload(BaseModel):
    site: str
    date: str
    red: bool


def weather_code_text(code) -> str:
    try:
        return WEATHER_CODE_TEXT.get(int(code), "Forecast unavailable")
    except (TypeError, ValueError):
        return "Forecast unavailable"


def format_temp(value) -> str:
    try:
        return f"{round(float(value))}°C"
    except (TypeError, ValueError):
        return "—"


def format_clock(dt: datetime) -> str:
    # Cross-platform 12-hour clock without a leading zero (avoids relying on
    # the non-portable "%-I" strftime directive).
    hour12 = dt.hour % 12 or 12
    return f"{hour12}:{dt.minute:02d} {'AM' if dt.hour < 12 else 'PM'}"


def hour_label(dt: datetime) -> str:
    """Plain-language hour label without minutes, e.g. '7 AM', '2 PM'."""
    hour12 = dt.hour % 12 or 12
    return f"{hour12} {'AM' if dt.hour < 12 else 'PM'}"


def chance_bucket(rain_chance) -> str:
    if rain_chance is None:
        return "none"
    if rain_chance < 15:
        return "none"
    if rain_chance < 40:
        return "low"
    if rain_chance < 60:
        return "medium"
    if rain_chance < 80:
        return "medium_high"
    return "high"


INTENSITY_BY_SEVERITY = {
    "green": "light",
    "yellow": "light to moderate",
    "orange": "moderate to heavy",
    "red": "heavy to intense",
}


def analyze_rain_day(hours: list[dict]) -> dict:
    """Characterize a day's hourly rain data for the plain-language sentence:
    whether rain covers most of the day, the main contiguous rain block, and
    whether a shorter, heavier peak stands out within that block."""
    broad = [
        h for h in hours
        if (h["prob"] or 0) >= BROAD_PROB_THRESHOLD or (h["precip"] or 0) >= BROAD_PRECIP_THRESHOLD
    ]
    if not broad:
        return {"has_rain": False}

    runs: list[list[dict]] = []
    current = [broad[0]]
    for h in broad[1:]:
        if h["time"] - current[-1]["time"] == timedelta(hours=1):
            current.append(h)
        else:
            runs.append(current)
            current = [h]
    runs.append(current)

    main_run = max(runs, key=lambda run: (sum(h["precip"] or 0 for h in run), len(run)))
    main_start = main_run[0]["time"]
    main_end = main_run[-1]["time"] + timedelta(hours=1)
    coverage_hours = sum(len(run) for run in runs)
    avg_precip = sum(h["precip"] or 0 for h in broad) / len(broad)

    heavy_hours = [h for h in main_run if (h["precip"] or 0) >= HEAVY_HOUR_PRECIP]
    has_peak = 0 < len(heavy_hours) < len(main_run)
    peak_start = heavy_hours[0]["time"] if has_peak else None
    peak_end = heavy_hours[-1]["time"] + timedelta(hours=1) if has_peak else None

    return {
        "has_rain": True,
        "main_start": main_start,
        "main_end": main_end,
        "coverage_full_day": coverage_hours >= FULL_DAY_HOURS,
        "avg_precip": avg_precip,
        "has_peak": has_peak,
        "peak_start": peak_start,
        "peak_end": peak_end,
    }


def build_narrative(sky_phrase: str, rain_chance, severity: str, analysis: dict) -> str:
    if not analysis.get("has_rain") or severity in (None, "none"):
        return f"{sky_phrase} skies throughout the day."

    intensity = INTENSITY_BY_SEVERITY.get(severity, "light")

    bucket = chance_bucket(rain_chance)
    if bucket == "none":
        bucket = "medium"  # rain is measurably happening; avoid contradicting that with "no chance"
    chance_phrase = CHANCE_PHRASES[bucket]

    coverage_full_day = analysis["coverage_full_day"]
    has_peak = analysis["has_peak"]

    if coverage_full_day and intensity == "moderate to heavy" and not has_peak:
        return f"Expect {chance_phrase} of consistent {intensity} rain throughout the day."

    if coverage_full_day:
        timing_clause = " throughout the day"
        if has_peak:
            timing_clause += f", particularly from {hour_label(analysis['peak_start'])} to {hour_label(analysis['peak_end'])}"
    else:
        main_start, main_end = analysis["main_start"], analysis["main_end"]
        if main_end.date() > main_start.date():
            timing_clause = f" from {hour_label(main_start)} until end of day"
        elif main_start.hour == 0:
            timing_clause = f" until {hour_label(main_end)}"
        else:
            timing_clause = f" from {hour_label(main_start)} to {hour_label(main_end)}"

    return f"{sky_phrase} skies with {chance_phrase} of {intensity} rain{timing_clause}."


def rain_window(hours: list[dict]) -> tuple[str, bool, float]:
    """Find the most likely contiguous rain period within one calendar day.

    ``hours`` is a list of {"time": datetime, "prob": float|None, "precip": float|None}
    sorted by time for a single date. Returns (label, has_window, max_hourly_precip_mm).
    """
    max_precip = max((h["precip"] or 0 for h in hours), default=0.0)
    flagged = [
        h for h in hours
        if (h["prob"] or 0) >= RAIN_HOUR_PROB_THRESHOLD or (h["precip"] or 0) >= RAIN_HOUR_PRECIP_THRESHOLD
    ]
    if not flagged:
        return "No significant rain expected", False, max_precip

    runs: list[list[dict]] = []
    current = [flagged[0]]
    for h in flagged[1:]:
        if h["time"] - current[-1]["time"] == timedelta(hours=1):
            current.append(h)
        else:
            runs.append(current)
            current = [h]
    runs.append(current)

    best = max(runs, key=lambda run: (sum(h["precip"] or 0 for h in run), len(run)))
    start = best[0]["time"]
    end = best[-1]["time"] + timedelta(hours=1)
    label = f"Rain likely {format_clock(start)} \u2013 {format_clock(end)}"
    return label, True, max_precip


def classify_severity(max_hourly_precip_mm: float, rain_chance, day_codes: list[int]) -> str:
    """Rain-intensity classification using standard hourly-rate thresholds
    (light < 2.5 mm/hr, moderate 2.5-7.5 mm/hr, heavy > 7.5 mm/hr), with any
    thunderstorm code forcing at least a heavy rating once rain is measurable."""
    has_thunder = any(code in THUNDER_CODES for code in day_codes)
    if max_hourly_precip_mm >= 7.5 or (has_thunder and max_hourly_precip_mm >= 2.5):
        return "orange"
    if max_hourly_precip_mm >= 2.5:
        return "yellow"
    if max_hourly_precip_mm >= 0.2 or (rain_chance or 0) >= 40 or has_thunder:
        return "green"
    return "none"


def fetch_all_forecasts() -> list[dict]:
    order = [name for _, name in SITES]
    lats = ",".join(str(SITE_COORDS[name][0]) for name in order)
    lons = ",".join(str(SITE_COORDS[name][1]) for name in order)
    params = {
        "latitude": lats,
        "longitude": lons,
        "daily": "weather_code,temperature_2m_max,temperature_2m_min,precipitation_sum,precipitation_probability_max,wind_gusts_10m_max",
        "hourly": "precipitation_probability,precipitation,weather_code,temperature_2m",
        "timezone": "Asia/Manila",
        "forecast_days": 5,
        "wind_speed_unit": "kmh",
    }
    response = requests.get(OPEN_METEO_FORECAST, params=params, timeout=25, headers={"User-Agent": "Open-Meteo-Weather-Tool/2.0"})
    response.raise_for_status()
    data = response.json()
    results = data if isinstance(data, list) else [data]
    if len(results) != len(order):
        raise HTTPException(502, "Open-Meteo returned an unexpected number of locations.")
    return results


def build_site_days(payload: dict) -> list[dict]:
    daily = payload.get("daily", {})
    hourly = payload.get("hourly", {})
    hourly_times = hourly.get("time", [])
    hourly_prob = hourly.get("precipitation_probability", [])
    hourly_precip = hourly.get("precipitation", [])
    hourly_code = hourly.get("weather_code", [])

    hours_by_date: dict[str, list[dict]] = {}
    for i, time_text in enumerate(hourly_times):
        date_key = time_text[:10]
        hours_by_date.setdefault(date_key, []).append({
            "time": datetime.fromisoformat(time_text),
            "prob": hourly_prob[i] if i < len(hourly_prob) else None,
            "precip": hourly_precip[i] if i < len(hourly_precip) else None,
            "code": hourly_code[i] if i < len(hourly_code) else None,
        })

    days = []
    dates = daily.get("time", [])
    for i, date_text in enumerate(dates):
        label = datetime.strptime(date_text, "%Y-%m-%d").strftime("%A %B %d, %Y")
        hours = sorted(hours_by_date.get(date_text, []), key=lambda h: h["time"])
        window_label, has_window, max_precip = rain_window(hours)
        day_codes = [h["code"] for h in hours if h["code"] is not None]
        code = daily.get("weather_code", [None] * len(dates))[i]
        rain_chance = daily.get("precipitation_probability_max", [None] * len(dates))[i]
        severity = classify_severity(max_precip, rain_chance, day_codes)
        gust = daily.get("wind_gusts_10m_max", [None] * len(dates))[i]
        sky_phrase = SKY_PHRASES.get(code, "Partly cloudy") if code is not None else "Partly cloudy"
        narrative = build_narrative(sky_phrase, rain_chance, severity, analyze_rain_day(hours))

        alert = ""
        alert_level = "none"
        overlay_source = ""
        if severity == "orange":
            alert = "Heavy rain modeled at its peak hour"
            alert_level = "heavy-rain"
            overlay_source = "Open-Meteo hourly model"
        elif any(c in THUNDER_CODES for c in day_codes):
            alert = "Thunderstorm risk somewhere in the day"
            alert_level = "thunderstorm"
            overlay_source = "Open-Meteo hourly model"
        if gust is not None and gust >= 62:
            gust_note = "Strong wind gusts possible"
            alert = f"{alert}; {gust_note}" if alert else gust_note
            alert_level = "wind" if alert_level == "none" else alert_level
            overlay_source = overlay_source or "Open-Meteo hourly model"

        days.append({
            "date": label,
            "condition": weather_code_text(code),
            "low": format_temp(daily.get("temperature_2m_min", [None] * len(dates))[i]),
            "high": format_temp(daily.get("temperature_2m_max", [None] * len(dates))[i]),
            "rain_chance": int(rain_chance) if rain_chance is not None else None,
            "rain_mm": round(max_precip, 1),
            "gust_kmh": round(gust, 1) if gust is not None else None,
            "icon": None,
            "forecast_window": window_label,
            "has_window": has_window,
            "narrative": narrative,
            "base_severity": severity,
            "automatic_severity": severity,
            "weather_alert": alert,
            "alert_level": alert_level,
            "overlay_source": overlay_source,
        })
    return days


def build_outlook_summary(rows: list[dict]) -> dict:
    candidates = [
        (row["site"], day) for row in rows for day in row["days"]
        if day.get("rain_mm", 0) or day.get("weather_alert")
    ]
    if not candidates:
        return {
            "available": True,
            "issued": f"Model run retrieved {datetime.now().astimezone():%Y-%m-%d %H:%M %Z}",
            "valid_until": "",
            "summary": "No significant rainfall or wind hazard is currently forecast at any monitored site in the next 5 days.",
            "source_url": OPEN_METEO_ATTRIBUTION,
        }
    ranked = sorted(candidates, key=lambda item: item[1].get("rain_mm", 0), reverse=True)[:5]
    sentences = []
    for site, day in ranked:
        detail = f"{site} on {day['date']}: {day['narrative']}"
        if day.get("weather_alert"):
            detail += f" \u2014 {day['weather_alert']}."
        sentences.append(detail)
    return {
        "available": True,
        "issued": f"Model run retrieved {datetime.now().astimezone():%Y-%m-%d %H:%M %Z}",
        "valid_until": "",
        "summary": " ".join(sentences),
        "source_url": OPEN_METEO_ATTRIBUTION,
    }


def fetch_current_situation() -> dict:
    """Tropical-cyclone status is not modeled by Open-Meteo, so this one feed
    still reads PAGASA's official daily bulletin, which is the authoritative
    source for named-storm advisories in the Philippines."""
    try:
        response = requests.get(PAGASA_DAILY_URL, timeout=25, headers={"User-Agent": "Mozilla/5.0 Weather-Tool/2.0"})
        response.raise_for_status()
        html = response.content.decode("utf-8", errors="replace")
        text = " ".join(BeautifulSoup(html, "html.parser").get_text(" ", strip=True).replace("\xa0", " ").replace("\ufffd", "\u00b0").split())

        def value(label: str, following: str) -> str:
            match = re.search(rf"{label}:?\s*(.+?)(?=\s+(?:{following}))", text, re.I)
            return match.group(1).strip() if match else "Unavailable"

        issued = re.search(r"Issued at:\s*(.+?)(?=\s+Synopsis)", text, re.I)
        synopsis = value("Synopsis", "TC Information|Forecast Weather Conditions")
        tc_block = re.search(r"TC Information\s+(.+?)(?=\s+Forecast Weather Conditions)", text, re.I)
        tc_text = tc_block.group(1).strip() if tc_block else "No tropical cyclone information published in the daily forecast."
        name_match = re.search(r"((?:SUPER\s+)?TYPHOON|TROPICAL STORM|SEVERE TROPICAL STORM|TROPICAL DEPRESSION)\s+([A-Z0-9() -]+?)(?=\s+LOCATION:)", tc_text, re.I)
        return {
            "available": True,
            "issued": issued.group(1).strip() if issued else "Issue time unavailable",
            "synopsis": synopsis,
            "tc_status": re.sub(r"\s+", " ", tc_text),
            "tc_name": " ".join(name_match.groups()).title() if name_match else "No named cyclone in daily forecast",
            "location": value("LOCATION", "MAXIMUM SUSTAINED WINDS"),
            "winds": value("MAXIMUM SUSTAINED WINDS", "GUSTINESS"),
            "gustiness": value("GUSTINESS", "MOVEMENT"),
            "movement": value("MOVEMENT", "Forecast Weather Conditions"),
            "source_url": PAGASA_DAILY_URL,
            "map_url": PANAHON_URL,
            "retrieved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        }
    except requests.RequestException as exc:
        return {"available": False, "error": str(exc), "source_url": PAGASA_DAILY_URL, "map_url": PANAHON_URL}


def fetch_historical_summary(site: str, month: int) -> dict:
    if site not in SITE_COORDS:
        raise HTTPException(400, "Unknown site.")
    if month < 1 or month > 12:
        raise HTTPException(400, "Month must be between 1 and 12.")

    now = datetime.now().astimezone()
    newest_year = now.year if month <= now.month else now.year - 1
    years = list(range(newest_year - 4, newest_year + 1))
    start_date = f"{years[0]}-{month:02d}-01"
    final_day = monthrange(years[-1], month)[1]
    if years[-1] == now.year and month == now.month:
        final_day = max(1, min(final_day, now.day - 5))
    end_date = f"{years[-1]}-{month:02d}-{final_day:02d}"
    latitude, longitude = SITE_COORDS[site]
    params = {
        "latitude": latitude,
        "longitude": longitude,
        "start_date": start_date,
        "end_date": end_date,
        "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,wind_gusts_10m_max,weather_code",
        "timezone": "Asia/Manila",
        "wind_speed_unit": "kmh",
        "precipitation_unit": "mm",
    }
    try:
        response = requests.get(OPEN_METEO_ARCHIVE, params=params, timeout=35, headers={"User-Agent": "Open-Meteo-Weather-Tool/2.0"})
        response.raise_for_status()
        daily = response.json().get("daily", {})
    except (requests.RequestException, ValueError) as exc:
        raise HTTPException(503, f"Historical weather source is temporarily unavailable: {exc}")

    records = []
    keys = ["time", "temperature_2m_max", "temperature_2m_min", "precipitation_sum", "wind_gusts_10m_max", "weather_code"]
    arrays = [daily.get(key, []) for key in keys]
    for values in zip(*arrays):
        date_text, high, low, rain, gust, code = values
        date = datetime.strptime(date_text, "%Y-%m-%d")
        if date.month == month and date.year in years:
            records.append({"date": date_text, "year": date.year, "high": high, "low": low, "rain": rain or 0, "gust": gust or 0, "code": code})

    summaries = []
    for year in years:
        items = [item for item in records if item["year"] == year]
        if not items:
            summaries.append({"year": year, "available": False})
            continue
        rainiest = max(items, key=lambda item: item["rain"])
        windiest = max(items, key=lambda item: item["gust"])
        critical = sorted(
            [item for item in items if item["rain"] >= 50 or item["gust"] >= 62 or item["code"] in (95, 96, 99)],
            key=lambda item: max(item["rain"] / 50, item["gust"] / 62), reverse=True,
        )[:6]
        if rainiest["rain"] >= 100:
            headline = "Extreme daily rainfall signal in the local reanalysis."
        elif rainiest["rain"] >= 50:
            headline = "At least one heavy-rain day in the local reanalysis."
        elif windiest["gust"] >= 62:
            headline = "At least one strong-gust day in the local reanalysis."
        else:
            headline = "No day crossed the dashboard's heavy-rain or strong-gust screening threshold."
        summaries.append({
            "year": year, "available": True, "days": len(items),
            "partial": year == now.year and month == now.month,
            "total_rain": round(sum(item["rain"] for item in items), 1),
            "wet_days": sum(item["rain"] >= 1 for item in items),
            "heavy_days": sum(item["rain"] >= 50 for item in items),
            "avg_high": round(sum(item["high"] for item in items) / len(items), 1),
            "avg_low": round(sum(item["low"] for item in items) / len(items), 1),
            "max_daily_rain": round(rainiest["rain"], 1), "rainiest_date": rainiest["date"],
            "max_gust": round(windiest["gust"], 1), "windiest_date": windiest["date"],
            "critical_days": critical, "headline": headline,
        })
    return {
        "site": site, "month": month, "month_name": datetime(2000, month, 1).strftime("%B"),
        "years": years, "summaries": summaries, "latitude": latitude, "longitude": longitude,
        "source": "Open-Meteo Historical Weather API (ERA5 reanalysis)",
        "source_url": "https://open-meteo.com/en/docs/historical-weather-api",
        "pagasa_annual_url": "https://www.pagasa.dost.gov.ph/tropical-cyclone/publications/annual-report",
        "pagasa_preliminary_url": "https://www.pagasa.dost.gov.ph/tropical-cyclone/publications/preliminary-report",
        "retrieved_at": now.isoformat(timespec="seconds"),
        "note": "Reanalysis identifies local weather signals, not cyclone causation. Verify named tropical cyclones in PAGASA reports.",
    }


def override_key(site: str, date: str) -> str:
    return f"{site.strip().casefold()}|{date.strip().casefold()}"


def _local_overrides() -> dict:
    try:
        return json.loads(LOCAL_OVERRIDES.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def load_overrides(force: bool = False) -> dict:
    now = time.time()
    if not force and now - _override_cache["loaded_at"] < 10:
        return dict(_override_cache["data"])
    if not os.getenv("BLOB_READ_WRITE_TOKEN"):
        data = _local_overrides()
    else:
        try:
            from vercel.blob import list_objects

            result = list_objects(prefix=OVERRIDE_PREFIX, limit=100)
            latest = max(result.blobs, key=lambda item: item.uploaded_at, default=None)
            data = requests.get(latest.url, params={"v": int(now)}, timeout=10).json() if latest else {}
        except Exception:
            data = dict(_override_cache["data"])
    _override_cache.update({"loaded_at": now, "data": data})
    return dict(data)


def save_overrides(data: dict) -> None:
    if os.getenv("BLOB_READ_WRITE_TOKEN"):
        from vercel.blob import BlobClient

        filename = f"{OVERRIDE_PREFIX}{int(time.time() * 1000)}-{secrets.token_hex(4)}.json"
        BlobClient().put(
            filename,
            json.dumps(data, separators=(",", ":")).encode("utf-8"),
            access="public",
            content_type="application/json",
            cache_control_max_age=60,
        )
    else:
        LOCAL_OVERRIDES.write_text(json.dumps(data, indent=2), encoding="utf-8")
    _override_cache.update({"loaded_at": time.time(), "data": dict(data)})


def verify_password(password: str) -> bool:
    encoded = os.getenv("ADMIN_PASSWORD_HASH", "")
    try:
        iterations_text, salt_hex, expected_hex = encoded.split("$", 2)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt_hex), int(iterations_text))
        return hmac.compare_digest(actual.hex(), expected_hex)
    except (TypeError, ValueError):
        return False


def session_token() -> str:
    expires = int(time.time()) + 8 * 60 * 60
    secret = os.getenv("ADMIN_SESSION_SECRET", "")
    signature = hmac.new(secret.encode(), str(expires).encode(), hashlib.sha256).hexdigest()
    return f"{expires}.{signature}"


def is_admin(request: Request) -> bool:
    token = request.cookies.get("pagasa_admin", "")
    secret = os.getenv("ADMIN_SESSION_SECRET", "")
    try:
        expires_text, signature = token.split(".", 1)
        expected = hmac.new(secret.encode(), expires_text.encode(), hashlib.sha256).hexdigest()
        return bool(secret) and int(expires_text) > int(time.time()) and hmac.compare_digest(signature, expected)
    except (TypeError, ValueError):
        return False


def build_payload() -> dict:
    forecasts = fetch_all_forecasts()
    overrides = load_overrides()
    rows = []
    for (region, site), site_payload in zip(SITES, forecasts):
        days = build_site_days(site_payload)
        for day in days:
            red = bool(overrides.get(override_key(site, day["date"])))
            day.update({
                "severity": "red" if red else day["automatic_severity"],
                "red_override": red,
                "severity_basis": "Admin override" if red else (day["overlay_source"] or "Open-Meteo hourly model"),
            })
        rows.append({
            "region": region,
            "site": site,
            "source_city": site,
            "days": days,
            "available": bool(days),
        })
    weekly = build_outlook_summary(rows)
    return {
        "issued": f"Open-Meteo model run retrieved {datetime.now().astimezone():%Y-%m-%d %H:%M %Z}",
        "retrieved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_mode": "live",
        "source_url": OPEN_METEO_ATTRIBUTION,
        "weekly_outlook": weekly,
        "rows": rows,
    }


def forecast_sentence(day: dict) -> str:
    """Plain-language sentence for one Excel cell, e.g. 'Partly cloudy skies
    with a high chance of light rain from 7 AM to 11 PM.' The severity color
    fill (green/yellow/orange/red) carries the risk level, so the text itself
    stays natural and readable rather than repeating a GREEN/YELLOW/ORANGE tag."""
    sentence = day.get("narrative") or f"{day['condition']}."
    if day.get("severity") == "red":
        sentence += " Admin-flagged: treat as elevated risk regardless of the modeled rain chance."
    return sentence


def export_workbook(payload: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "5-Day Forecast"
    ws.sheet_view.showGridLines = False
    dates = []
    for row in payload["rows"]:
        if row["days"]:
            dates = [day["date"] for day in row["days"][:5]]
            break
    while len(dates) < 5:
        dates.append(f"Day {len(dates) + 1}")

    ws.append([None, "Site", *dates])
    for row in payload["rows"]:
        values = [row["region"], row["site"]]
        values += [forecast_sentence(d) for d in row["days"][:5]]
        values += ["Forecast unavailable"] * (7 - len(values))
        ws.append(values)

    severity_fills = {
        "green": PatternFill("solid", fgColor="00B050"),
        "yellow": PatternFill("solid", fgColor="FFFF00"),
        "orange": PatternFill("solid", fgColor="FFC000"),
        "red": PatternFill("solid", fgColor="FF0000"),
        "none": PatternFill("solid", fgColor="BFBFBF"),
    }
    severity_fonts = {"green": "FFFFFF", "yellow": "3D2E00", "orange": "3D2E00", "red": "FFFFFF", "none": "404040"}
    for row_index, row in enumerate(payload["rows"], start=2):
        for day_index, day in enumerate(row["days"][:5], start=3):
            # Use the final displayed severity. Admin overrides set this value to red.
            final_severity = day.get("severity", "none")
            cell = ws.cell(row_index, day_index)
            cell.fill = severity_fills.get(final_severity, severity_fills["none"])
            cell.font = Font(color=severity_fonts.get(final_severity, severity_fonts["none"]), bold=final_severity in ("red", "orange"))

    navy, blue, white = "17365D", "D9EAF7", "FFFFFF"
    ws["B1"].fill = PatternFill("solid", fgColor=navy)
    for cell in ws[1][1:7]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A6A6A6")
    for row in ws.iter_rows(min_row=2, max_row=13, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].fill = PatternFill("solid", fgColor=navy)
        row[0].font = Font(color=white, bold=True)
        row[1].fill = PatternFill("solid", fgColor=blue)
        row[1].font = Font(bold=True)
    ws.merge_cells("A2:A8")
    ws.merge_cells("A9:A10")
    ws.merge_cells("A11:A13")
    for cell in (ws["A2"], ws["A9"], ws["A11"]):
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
    widths = [13, 18, 38, 38, 38, 38, 38]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34
    for i in range(2, 14):
        ws.row_dimensions[i].height = 84
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "B1:G13"

    meta = wb.create_sheet("Source")
    meta.append(["Field", "Value"])
    meta.append(["Forecast source", "Open-Meteo (open-meteo.com)"])
    meta.append(["Issued", payload["issued"]])
    meta.append(["Retrieved", payload["retrieved_at"]])
    meta.append(["Source mode", payload["source_mode"]])
    meta.append([])
    meta.append(["Rainfall intensity", "Meaning"])
    legend = [
        ("GREEN", "Light rain (<2.5 mm/hr modeled)", "00B050", "FFFFFF"),
        ("YELLOW", "Moderate rain (2.5\u20137.5 mm/hr modeled)", "FFFF00", "3D2E00"),
        ("ORANGE", "Heavy rain (>7.5 mm/hr modeled) or thunderstorm risk", "FFC000", "3D2E00"),
        ("RED", "Admin override; discretionary escalation applied in production", "FF0000", "FFFFFF"),
        ("GRAY", "No rain classification", "BFBFBF", "404040"),
    ]
    for level, meaning, fill, font_color in legend:
        meta.append([level, meaning])
        for cell in meta[meta.max_row]:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font_color, bold=True)
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 95
    meta["A1"].font = meta["B1"].font = Font(bold=True, color=white)
    meta["A1"].fill = meta["B1"].fill = PatternFill("solid", fgColor=navy)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/")
def home():
    return FileResponse(STATIC / "index.html")


@app.get("/monitor")
def monitor():
    return FileResponse(STATIC / "monitor.html")


@app.get("/guidance")
def guidance():
    return FileResponse(STATIC / "guidance.html")


@app.get("/history")
def history():
    return FileResponse(STATIC / "history.html")


@app.get("/api/forecast")
def forecast():
    return JSONResponse(build_payload())


@app.get("/api/situation")
def situation():
    return JSONResponse(fetch_current_situation())


@app.get("/api/history")
def history_api(site: str = "Metro Manila", month: int = datetime.now().month):
    return JSONResponse(fetch_historical_summary(site, month))


@app.get("/api/admin/status")
def admin_status(request: Request):
    return {"authenticated": is_admin(request), "configured": bool(os.getenv("ADMIN_PASSWORD_HASH"))}


@app.post("/api/admin/login")
def admin_login(payload: LoginPayload, response: Response):
    if not os.getenv("ADMIN_PASSWORD_HASH"):
        raise HTTPException(503, "Admin access is not configured.")
    if not verify_password(payload.password):
        raise HTTPException(401, "Incorrect admin password.")
    response.set_cookie(
        "pagasa_admin",
        session_token(),
        max_age=8 * 60 * 60,
        httponly=True,
        secure=bool(os.getenv("VERCEL")),
        samesite="strict",
    )
    return {"authenticated": True}


@app.post("/api/admin/logout")
def admin_logout(response: Response):
    response.delete_cookie("pagasa_admin")
    return {"authenticated": False}


@app.put("/api/admin/override")
def update_override(payload: OverridePayload, request: Request):
    if not is_admin(request):
        raise HTTPException(401, "Admin login required.")
    valid_sites = {site.casefold() for _, site in SITES}
    if payload.site.casefold() not in valid_sites or not payload.date.strip():
        raise HTTPException(400, "Invalid site or forecast date.")
    data = load_overrides(force=True)
    key = override_key(payload.site, payload.date)
    if payload.red:
        data[key] = {"red": True, "updated_at": datetime.now().astimezone().isoformat(timespec="seconds")}
    else:
        data.pop(key, None)
    save_overrides(data)
    return {"site": payload.site, "date": payload.date, "red_override": payload.red}


@app.get("/api/export")
def export():
    payload = build_payload()
    stream = export_workbook(payload)
    filename = f"5-Day_Forecast_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
